from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook

from app.domain.allocation_rules import AllocationType
from app.domain.notification_types import NotificationType
from app.domain.work_profile import WORK_LOCATION_TYPE_VALUES, normalize_choice
from app.domain.email_templates import ONBOARD_INVITE
from app.domain.message_constants import ONBOARD_INVITE_SUBJECT
from app.repositories.designation_repository import DesignationRepository
from app.repositories.employee_repository import EmployeeRepository
from app.repositories.leave_repository import LeaveRepository
from app.repositories.profile_repository import ProfileRepository
from app.schemas.allocation import AllocationCreateRequest
from app.schemas.attrition import AttritionRecordResponse, AttritionUpsertRequest
from app.schemas.employee import (
    EmployeeProfileHrUpdate,
    EmployeeProfileResponse,
    OnboardListItem,
    OnboardListResponse,
    OnboardUserResponse,
    ProfileUpdateRequest,
    UserOnboardCreate,
    UserOnboardUpdate,
)
from app.services.notification_service import NotificationService
from app.services.email_service import EmailService
from app.services.reporting_service import ReportingService


class EmployeeService:
    _IMPORT_HEADER_ALIASES: dict[str, set[str]] = {
        "emp_id": {"emp_id", "employee_id", "empid", "emp id"},
        "email": {"email", "email_id", "user_email", "mail"},
        "name": {"name", "employee_name", "full_name"},
        "status": {"status", "user_status"},
        "user_type": {"user_type", "employment_type", "type"},
        "department": {"department", "stream"},
        "phone_number": {"phone_number", "phone", "phone_no", "mobile", "mobile_number"},
        "role": {"role", "designation", "title"},
        "work_mode": {"work_mode", "workmode", "mode"},
        "work_location_type": {"work_location_type", "work_location", "location_type"},
        "doj": {"doj", "date_of_joining", "joining_date"},
        "doi": {"doi", "date_of_internship", "internship_start_date"},
        "internship_duration": {"internship_duration", "internship_duration_months", "duration"},
        "band_id": {"band_id", "band", "bandid"},
    }
    _USER_DATA_LEGACY_INDEX = {
        "email": 1,
        "user_type": 2,
        "status": 3,
        "band_id": 4,
        "doj": 5,
        "phone_number": 6,
        "department": 7,
    }
    _USER_BATCH_LEGACY_INDEX = {
        "emp_id": 0,
        "name": 1,
        "email": 2,
    }

    def __init__(self, db) -> None:
        self.db = db
        self.employee_repo = EmployeeRepository(db)
        self.designation_repo = DesignationRepository(db)
        self.profile_repo = ProfileRepository(db)
        self.leave_repo = LeaveRepository(db)
        self.notification_service = NotificationService(db)
        self.email_service = EmailService()

    def _to_file_url(self, file: UploadFile) -> str:
        timestamp = int(datetime.now(UTC).timestamp())
        return f"local://uploads/{timestamp}_{file.filename}"

    @staticmethod
    def _as_utc_datetime(value: date | None) -> datetime | None:
        if value is None:
            return None
        return datetime(value.year, value.month, value.day, tzinfo=UTC)

    def _validate_primary_skills(self, skills: list[str]) -> list[str]:
        clean_skills = [skill.strip() for skill in skills if skill and skill.strip()]
        if not clean_skills:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="At least one Primary Skill is required",
            )
        return clean_skills

    def _initial_leave_balance(self, user_type: str, join_date: date | None) -> tuple[float, float]:
        if user_type.upper() == "INTERN":
            return 0.0, 0.0

        # Initial leave must be derived from DOJ, not record creation date.
        if not join_date:
            return 0.0, 0.0

        if join_date.day >= 15:
            return 0.0, 0.0

        if join_date.month < 8:
            return 1.5, 0.0

        if join_date.month >= 8:
            return 0.0, 1.5
        return 0.0, 0.0

    async def _send_onboarding_invite(self, email: str, name: str) -> None:
        # Java-parity onboarding invite email.
        subject = ONBOARD_INVITE_SUBJECT
        body = ONBOARD_INVITE % name
        await self.email_service.send_email(to=email, subject=subject, body=body, cc=None, is_html=True)

    @staticmethod
    def _normalize_header(value: Any) -> str:
        if value is None:
            return ""
        normalized = str(value).strip().lower()
        for ch in (" ", "-", ".", "/"):
            normalized = normalized.replace(ch, "_")
        while "__" in normalized:
            normalized = normalized.replace("__", "_")
        return normalized.strip("_")

    def _build_header_map(self, first_row: tuple[Any, ...] | list[Any] | None) -> tuple[dict[str, int], bool]:
        if not first_row:
            return {}, False
        alias_to_field: dict[str, str] = {}
        for field, aliases in self._IMPORT_HEADER_ALIASES.items():
            for alias in aliases:
                alias_to_field[alias] = field
        mapping: dict[str, int] = {}
        recognized = 0
        for idx, cell in enumerate(first_row):
            token = self._normalize_header(cell)
            if not token:
                continue
            field = alias_to_field.get(token)
            if field and field not in mapping:
                mapping[field] = idx
                recognized += 1
        has_header = recognized >= 2
        return (mapping if has_header else {}), has_header

    @staticmethod
    def _as_clean_text(value: Any) -> str | None:
        if value is None:
            return None
        txt = str(value).strip()
        return txt or None

    @staticmethod
    def _parse_date_cell(value: Any) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        txt = str(value).strip()
        if not txt:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(txt, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(txt).date()
        except ValueError:
            return None

    @staticmethod
    def _parse_int_cell(value: Any) -> int | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        txt = str(value).strip()
        if not txt:
            return None
        try:
            return int(float(txt))
        except ValueError:
            return None

    @staticmethod
    def _normalize_work_location_type(value: Any) -> str | None:
        try:
            return normalize_choice(EmployeeService._as_clean_text(value), WORK_LOCATION_TYPE_VALUES, "work_location_type")
        except ValueError:
            return None

    def _normalize_user_import_row(self, row: tuple[Any, ...], header_map: dict[str, int], source: str) -> dict[str, Any]:
        raw: dict[str, Any] = {}
        for field, idx in header_map.items():
            if idx < len(row):
                raw[field] = row[idx]

        legacy_map = self._USER_BATCH_LEGACY_INDEX if source == "batch" else self._USER_DATA_LEGACY_INDEX
        for field, idx in legacy_map.items():
            if raw.get(field) is None and idx < len(row):
                raw[field] = row[idx]

        out: dict[str, Any] = {}
        emp_id = self._as_clean_text(raw.get("emp_id"))
        if emp_id:
            out["emp_id"] = emp_id

        email = self._as_clean_text(raw.get("email"))
        if email:
            out["email"] = email.lower()

        name = self._as_clean_text(raw.get("name"))
        if name:
            out["name"] = name

        status = self._as_clean_text(raw.get("status"))
        if status:
            out["status"] = status.upper()

        user_type = self._as_clean_text(raw.get("user_type"))
        if user_type:
            out["user_type"] = user_type.upper()

        department = self._as_clean_text(raw.get("department"))
        if department:
            out["department"] = department

        phone_number = self._as_clean_text(raw.get("phone_number"))
        if phone_number:
            out["phone_number"] = phone_number

        role = self._as_clean_text(raw.get("role"))
        if role:
            out["role"] = role

        work_mode = self._as_clean_text(raw.get("work_mode"))
        if work_mode:
            out["work_mode"] = work_mode.upper()

        work_location_type = self._normalize_work_location_type(raw.get("work_location_type"))
        if work_location_type:
            out["work_location_type"] = work_location_type

        doj = self._parse_date_cell(raw.get("doj"))
        if doj:
            out["doj"] = doj

        doi = self._parse_date_cell(raw.get("doi"))
        if doi:
            out["doi"] = doi

        internship_duration = self._parse_int_cell(raw.get("internship_duration"))
        if internship_duration is not None:
            out["internship_duration"] = internship_duration

        band_id = self._parse_int_cell(raw.get("band_id"))
        if band_id is not None:
            out["band_id"] = band_id

        return out

    @staticmethod
    def _to_repo_payload(user_fields: dict[str, Any], *, include_email: bool = False) -> dict[str, Any]:
        payload: dict[str, Any] = {}
        if user_fields.get("emp_id") is not None:
            payload["empId"] = user_fields["emp_id"]
        if include_email and user_fields.get("email"):
            payload["email"] = user_fields["email"]
        if user_fields.get("name") is not None:
            payload["name"] = user_fields["name"]
        if user_fields.get("status") is not None:
            payload["status"] = user_fields["status"]
        if user_fields.get("user_type") is not None:
            payload["userType"] = user_fields["user_type"]
        if user_fields.get("department") is not None:
            payload["department"] = user_fields["department"]
        if user_fields.get("phone_number") is not None:
            payload["phoneNumber"] = user_fields["phone_number"]
        if user_fields.get("role") is not None:
            payload["role"] = user_fields["role"]
        if user_fields.get("work_mode") is not None:
            payload["workMode"] = user_fields["work_mode"]
        if user_fields.get("work_location_type") is not None:
            payload["workLocationType"] = user_fields["work_location_type"]
        if user_fields.get("doj") is not None:
            payload["doj"] = user_fields["doj"]
        if user_fields.get("doi") is not None:
            payload["doi"] = user_fields["doi"]
        if user_fields.get("internship_duration") is not None:
            payload["internshipDuration"] = user_fields["internship_duration"]
        if user_fields.get("band_id") is not None:
            payload["bandId"] = user_fields["band_id"]
        return payload

    async def create_user_onboard(self, payload: UserOnboardCreate) -> OnboardUserResponse:
        if await self.employee_repo.get_user_by_email(payload.email):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="User with email already exists")
        if payload.band_id and not await self.employee_repo.get_band(payload.band_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Band not found with id : {payload.band_id}")
        if payload.role and payload.band_id and payload.department:
            designation_exists = await self.designation_repo.exists_by_band_department_and_name(
                payload.band_id,
                payload.department,
                payload.role,
            )
            if not designation_exists:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid designation for selected band and stream",
                )

        async with self.db.tx() as transaction:
            doj_dt = self._as_utc_datetime(payload.doj)
            doi_dt = self._as_utc_datetime(payload.doi)
            user = await self.employee_repo.create_user(
                {
                    "email": payload.email,
                    "name": payload.name,
                    "userType": payload.user_type,
                    "department": payload.department,
                    "deliveryStatus": payload.delivery_status,
                    "phoneNumber": payload.phone_number,
                    "workMode": payload.work_mode,
                    "workLocationType": payload.work_location_type,
                    "role": payload.role,
                    "bandId": payload.band_id,
                    "doj": doj_dt,
                    "doi": doi_dt,
                    "internshipDuration": payload.internship_duration,
                    "status": "INVITED",
                },
                client=transaction,
            )
            user = await self.employee_repo.update_user(user.id, {"empId": str(user.id)}, client=transaction)
            await self.profile_repo.get_or_create_profile(user.id, client=transaction)

            primary_leave, secondary_leave = self._initial_leave_balance(payload.user_type, payload.doj)
            today = date.today()
            await self.leave_repo.create_mapping(user.id, today.year, today.month, primary_leave, secondary_leave, client=transaction)

            employee_role = await self.employee_repo.get_or_create_role("ROLE_EMPLOYEE")
            await self.employee_repo.assign_role(user.id, employee_role.id, project_code=None, client=transaction)
            start_date = self._as_utc_datetime(today)
            await self.employee_repo.create_bench_allocation(user.id, payload.role, start_date, client=transaction)

        await self._send_onboarding_invite(user.email, user.name)
        await self.notification_service.send_notification(
            receiver_id=user.id,
            sender_id=None,
            notification_type=NotificationType.ONBOARDING_INVITE,
            title="Onboarding Started",
            message="Welcome to Webtrak. Please complete your onboarding profile.",
        )

        return OnboardUserResponse(
            emp_id=user.empId or str(user.id),
            email=user.email,
            name=user.name,
            status=user.status,
            user_type=user.userType,
            work_location_type=user.workLocationType,
        )

    async def update_user_onboard(
        self,
        payload: UserOnboardUpdate,
        resume: UploadFile | None,
        reliving_letter: UploadFile | None,
        salary_slips: list[UploadFile] | None,
        certifications: list[UploadFile] | None,
        profile_photo: UploadFile | None,
        aadhaar: UploadFile | None,
        pan_card: UploadFile | None,
    ) -> OnboardUserResponse:
        user = await self.profile_repo.get_user_by_email(payload.email)
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

        profile = await self.profile_repo.get_or_create_profile(user.id)
        profile_update = {
            "yoe": payload.yoe,
            "primarySkills": self._validate_primary_skills(payload.primary_skills),
            "secondarySkills": payload.secondary_skills,
        }

        if payload.yoe and payload.yoe > 0 and user.userType.upper() == "FULLTIME" and not payload.experience:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Please provide Experience details if you are experienced",
            )
        profile_update["experience"] = payload.experience

        if user.userType.upper() in {"FULLTIME", "INTERN"} and resume is None and not profile.personalResume:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Please upload your updated Resume")
        if profile_photo is None and not profile.profilePhoto:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Please upload profile photo")
        if aadhaar is None and not profile.aadhaar:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Please upload Aadhaar")
        if pan_card is None and not profile.panCard:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Please upload PAN card")

        async with self.db.tx() as transaction:
            if resume:
                resume_url = self._to_file_url(resume)
                profile_update["personalResume"] = resume_url
                await self.profile_repo.add_document(user.id, "RESUME", resume_url, client=transaction)
            if profile_photo:
                photo_url = self._to_file_url(profile_photo)
                profile_update["profilePhoto"] = photo_url
                await self.profile_repo.add_document(user.id, "PHOTO", photo_url, client=transaction)
            if aadhaar:
                aadhaar_url = self._to_file_url(aadhaar)
                profile_update["aadhaar"] = aadhaar_url
                await self.profile_repo.add_document(user.id, "AADHAAR", aadhaar_url, client=transaction)
            if pan_card:
                pan_url = self._to_file_url(pan_card)
                profile_update["panCard"] = pan_url
                await self.profile_repo.add_document(user.id, "PAN", pan_url, client=transaction)

            if payload.yoe and payload.yoe > 0:
                if reliving_letter:
                    await self.profile_repo.add_document(
                        user.id, "RELIVING_LETTER", self._to_file_url(reliving_letter), client=transaction
                    )
                if salary_slips:
                    for slip in salary_slips:
                        await self.profile_repo.add_document(
                            user.id, "SALARY_SLIP", self._to_file_url(slip), client=transaction
                        )
            if certifications:
                for certification in certifications:
                    await self.profile_repo.add_document(
                        user.id,
                        "CERTIFICATION",
                        self._to_file_url(certification),
                        client=transaction,
                    )

            await self.profile_repo.update_profile(user.id, profile_update, client=transaction)
            user_updates: dict[str, str] = {"status": "ACTIVE"}
            if payload.work_location_type is not None:
                user_updates["workLocationType"] = payload.work_location_type
            await self.profile_repo.update_user(user.id, user_updates, client=transaction)

        user = await self.profile_repo.get_user_by_email(payload.email)

        return OnboardUserResponse(
            emp_id=user.empId or str(user.id),
            email=user.email,
            name=user.name,
            status=user.status,
            user_type=user.userType,
            work_location_type=user.workLocationType,
        )

    async def get_profile_by_email(self, email: str) -> EmployeeProfileResponse:
        user = await self.profile_repo.get_user_by_email(email)
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
        profile = await self.profile_repo.get_or_create_profile(user.id)
        return EmployeeProfileResponse(
            emp_id=user.empId,
            email=user.email,
            name=user.name,
            status=user.status,
            user_type=user.userType,
            work_location_type=user.workLocationType,
            department=user.department,
            phone_number=user.phoneNumber,
            work_mode=user.workMode,
            yoe=profile.yoe,
            experience=profile.experience,
            primary_skills=profile.primarySkills or [],
            secondary_skills=profile.secondarySkills or [],
            profile_photo=profile.profilePhoto,
            personal_resume=profile.personalResume,
        )

    async def update_profile(self, email: str, payload: ProfileUpdateRequest, profile_pic: UploadFile | None) -> EmployeeProfileResponse:
        user = await self.profile_repo.get_user_by_email(email)
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

        user_update = {}
        profile_update = {}
        if payload.phone_number is not None:
            user_update["phoneNumber"] = payload.phone_number
        if payload.work_mode is not None:
            user_update["workMode"] = payload.work_mode
        if payload.work_location_type is not None:
            user_update["workLocationType"] = payload.work_location_type
        if payload.primary_skills is not None:
            profile_update["primarySkills"] = self._validate_primary_skills(payload.primary_skills)
        if payload.secondary_skills is not None:
            profile_update["secondarySkills"] = payload.secondary_skills
        if payload.experience is not None:
            profile_update["experience"] = payload.experience
        if profile_pic:
            profile_update["profilePhoto"] = self._to_file_url(profile_pic)

        async with self.db.tx() as transaction:
            if user_update:
                await self.profile_repo.update_user(user.id, user_update, client=transaction)
            if profile_update:
                await self.profile_repo.update_profile(user.id, profile_update, client=transaction)
        return await self.get_profile_by_email(email)

    async def get_employee_profile(self, emp_id: str) -> EmployeeProfileResponse:
        user = await self.profile_repo.get_user_by_emp_id(emp_id)
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
        return await self.get_profile_by_email(user.email)

    async def update_employee_profile(self, emp_id: str, payload: EmployeeProfileHrUpdate) -> EmployeeProfileResponse:
        user = await self.profile_repo.get_user_by_emp_id(emp_id)
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
        user_update = {}
        profile_update = {}

        if payload.name is not None:
            user_update["name"] = payload.name
        if payload.department is not None:
            user_update["department"] = payload.department
        if payload.user_status is not None:
            user_update["status"] = payload.user_status
        if payload.work_mode is not None:
            user_update["workMode"] = payload.work_mode
        if payload.work_location_type is not None:
            user_update["workLocationType"] = payload.work_location_type
        if payload.band_id is not None:
            if not await self.employee_repo.get_band(payload.band_id):
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Band not found with id : {payload.band_id}")
            user_update["bandId"] = payload.band_id
        if payload.primary_skills is not None:
            profile_update["primarySkills"] = self._validate_primary_skills(payload.primary_skills)
        if payload.secondary_skills is not None:
            profile_update["secondarySkills"] = payload.secondary_skills
        if payload.experience is not None:
            profile_update["experience"] = payload.experience
        if payload.yoe is not None:
            profile_update["yoe"] = payload.yoe

        async with self.db.tx() as transaction:
            if user_update:
                await self.profile_repo.update_user(user.id, user_update, client=transaction)
            if profile_update:
                await self.profile_repo.update_profile(user.id, profile_update, client=transaction)
        return await self.get_profile_by_email(user.email)

    async def get_onboarded_users(
        self,
        page: int,
        size: int,
        search: str | None,
        user_type: str | None,
        onboarding_status: str | None,
    ) -> OnboardListResponse:
        users, total = await self.employee_repo.list_onboard_users(page, size, search, user_type, onboarding_status)
        return OnboardListResponse(
            items=[
                OnboardListItem(
                    emp_id=user.empId,
                    email=user.email,
                    name=user.name,
                    status=user.status,
                    user_type=user.userType,
                    department=user.department,
                )
                for user in users
            ],
            total=total,
            page=page,
            size=size,
        )

    async def import_leave_data(self, content: bytes) -> dict[str, int | str]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        current_year = date.today().year
        current_month = date.today().month
        processed = 0
        skipped = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            email = str(row[0]).strip().lower() if row[0] is not None else ""
            if not email:
                continue
            primary = float(row[2] or 0.0)
            secondary = float(row[3] or 0.0)
            carry_forward = float(row[4] or 0.0)
            user = await self.employee_repo.get_user_by_email(email)
            if not user:
                skipped += 1
                continue
            async with self.db.tx() as tx:
                mapping = await self.leave_repo.get_mapping(user.id, current_year, current_month, client=tx)
                if mapping is None:
                    await self.leave_repo.create_mapping(
                        user.id,
                        current_year,
                        current_month,
                        primary,
                        secondary,
                        carry_forward=carry_forward,
                        client=tx,
                    )
                else:
                    mapping.primary_leave = primary
                    mapping.secondary_leave = secondary
                    mapping.carry_forward = carry_forward
                    await self.leave_repo.save_mapping(mapping, client=tx)
            processed += 1
        return {"processed": processed, "skipped": skipped, "message": "Leave data imported successfully"}

    async def import_user_data(self, content: bytes) -> dict[str, int | str]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        processed = 0
        skipped = 0
        errors: list[str] = []
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_map, has_header = self._build_header_map(first_row)
        start_row = 2 if has_header else 1
        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if not row:
                continue
            normalized = self._normalize_user_import_row(row, header_map, source="user_data")
            email = normalized.get("email", "")
            if not email:
                skipped += 1
                continue
            user = await self.employee_repo.get_user_by_email(email)
            if not user:
                skipped += 1
                continue
            payload = self._to_repo_payload(normalized, include_email=False)
            if payload:
                try:
                    await self.employee_repo.update_user(user.id, payload)
                except Exception as exc:
                    errors.append(f"row {row_idx}: {exc}")
                    skipped += 1
                    continue
                processed += 1
            else:
                skipped += 1
        message = "User data imported successfully"
        if errors:
            message = f"{message}; {len(errors)} row(s) skipped with errors"
        return {"processed": processed, "skipped": skipped, "message": message, "errors": errors[:50]}

    async def bulk_upload_users(self, content: bytes) -> dict[str, int | str]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        processed = 0
        skipped = 0
        errors: list[str] = []
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_map, has_header = self._build_header_map(first_row)
        start_row = 2 if has_header else 1
        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if not row:
                continue
            normalized = self._normalize_user_import_row(row, header_map, source="batch")
            email = normalized.get("email", "")
            if not email:
                skipped += 1
                continue
            existing = await self.employee_repo.get_user_by_email(email)
            async with self.db.tx() as tx:
                try:
                    if existing:
                        update_payload = self._to_repo_payload(normalized, include_email=False)
                        if update_payload:
                            await self.employee_repo.update_user(existing.id, update_payload, client=tx)
                        processed += 1
                        continue
                    create_payload = self._to_repo_payload(normalized, include_email=True)
                    if "name" not in create_payload or not str(create_payload["name"]).strip():
                        create_payload["name"] = email.split("@")[0]
                    create_payload.setdefault("status", "INVITED")
                    create_payload.setdefault("userType", "FULLTIME")
                    created = await self.employee_repo.create_user(create_payload, client=tx)
                    if not (getattr(created, "empId", None) or getattr(created, "emp_id", None)):
                        created = await self.employee_repo.update_user(created.id, {"empId": str(created.id)}, client=tx)
                    role = await self.employee_repo.get_or_create_role("ROLE_EMPLOYEE")
                    await self.employee_repo.assign_role(created.id, role.id, project_code="GLOBAL", client=tx)
                    created_doj = getattr(created, "doj", None)
                    if isinstance(created_doj, datetime):
                        created_doj = created_doj.date()
                    if not isinstance(created_doj, date):
                        created_doj = None
                    user_type = str(getattr(created, "userType", create_payload.get("userType", "FULLTIME")) or "FULLTIME")
                    primary_leave, secondary_leave = self._initial_leave_balance(user_type, created_doj)
                    today = date.today()
                    await self.leave_repo.create_mapping(
                        created.id,
                        today.year,
                        today.month,
                        primary_leave,
                        secondary_leave,
                        client=tx,
                    )
                    start_date = self._as_utc_datetime(today)
                    await self.employee_repo.create_bench_allocation(created.id, getattr(created, "role", None), start_date, client=tx)
                    processed += 1
                except Exception as exc:
                    errors.append(f"row {row_idx}: {exc}")
                    skipped += 1
        message = f"Import complete - processed {processed} records"
        if skipped:
            message = f"{message}; skipped {skipped} row(s)"
        return {"processed": processed, "skipped": skipped, "message": message, "errors": errors[:50]}

    async def import_allocations_legacy(self, content: bytes, actor_roles: set[str]) -> dict[str, int | str | list[str]]:
        from app.services.allocation_service import AllocationService

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        allocation_service = AllocationService(self.db)
        completed = 0
        errors: list[str] = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            email = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
            role = str(row[3]).strip() if len(row) > 3 and row[3] else "Employee"
            allocation_value = float(row[4]) if len(row) > 4 and row[4] is not None else 1.0
            start_date = row[5] if len(row) > 5 else None
            end_date = row[6] if len(row) > 6 else None
            locked_date = row[7] if len(row) > 7 else None
            allocation_type = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            project_code = str(row[9]).strip().upper() if len(row) > 9 and row[9] else ""
            if not email or not project_code or not start_date:
                continue
            try:
                if isinstance(start_date, datetime):
                    start_date = start_date.date()
                if isinstance(end_date, datetime):
                    end_date = end_date.date()
                if isinstance(locked_date, datetime):
                    locked_date = locked_date.date()

                mapped_type = AllocationType.DEPLOYABLE
                if allocation_type.upper() == "NB":
                    mapped_type = AllocationType.NONBILLABLE
                elif allocation_type.lower() == "staffing":
                    mapped_type = AllocationType.NONBILLABLE
                elif allocation_type.lower() == "locked":
                    mapped_type = AllocationType.LOCKED

                normalized_role = role
                if role == "Delivery Manager":
                    normalized_role = "DM"
                elif role == "Project Manager":
                    normalized_role = "PM"
                elif role == "Account Manager":
                    normalized_role = "AM"

                payload = AllocationCreateRequest(
                    employee_email=email,
                    project_code=project_code,
                    role=normalized_role,
                    allocated_hours=max(1, min(8, int(allocation_value * 8))),
                    start_date=start_date,
                    end_date=end_date if isinstance(end_date, date) else None,
                    allocation_type=mapped_type,
                    locked_in_date=locked_date if isinstance(locked_date, date) else None,
                    is_manager=normalized_role in {"PM", "DM", "AM"},
                )
                await allocation_service.add_allocation(payload, actor_roles=actor_roles)
                completed += 1
            except Exception as exc:
                errors.append(f"row {idx}: {exc}")
        message = "Allocation data imported successfully"
        if errors:
            message = f"{message}; {len(errors)} row(s) skipped"
        return {"completed": completed, "message": message, "errors": errors[:50]}

    async def offboard_employee(self, *, actor_email: str, emp_id: str, payload: AttritionUpsertRequest) -> AttritionRecordResponse:
        return await ReportingService(self.db).upsert_attrition(actor_email=actor_email, emp_id=emp_id, payload=payload)
